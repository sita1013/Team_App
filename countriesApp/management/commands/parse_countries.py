import os
from pathlib import Path
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from countriesApp.models import Country, Indicator, Measurement 

class Command(BaseCommand):
    help = 'Parse countries and air pollution data from Excel and store in the database'

    def handle(self, *args, **options):
        # Define path to the Excel file
        base_dir = Path(__file__).resolve().parent.parent.parent.parent
        file_name = 'pm25_data.xlsx'
        file_path = os.path.join(base_dir, 'countriesApp', 'countries_data', file_name)  # Adjust folder path if needed
        print(f"Looking for file at: {file_path}.") 

        # Load workbook and select the main data sheet
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook['Data']
        print(f'Opened sheet: {sheet.title}')

        # Optional: clear existing data
        Measurement.objects.all().delete()
        Indicator.objects.all().delete()
        Country.objects.all().delete()
        print("Existing data cleared.")

        # Get the year headers (row 4, index 3 in 0-based)
        header_row = 3
        years = [cell.value for cell in sheet[header_row][4:] if isinstance(cell.value, int)]

        # Process each data row
        for row in sheet.iter_rows(min_row=5):  # Skip headers
            country_name = row[0].value
            country_code = row[1].value
            indicator_name = row[2].value
            indicator_code = row[3].value

            if not country_name or not indicator_name:
                continue  # Skip empty rows

            # Create or get Country and Indicator
            country, _ = Country.objects.get_or_create(name=country_name, code=country_code)
            indicator, _ = Indicator.objects.get_or_create(name=indicator_name, code=indicator_code)

            # Loop through years and values
            for i, year in enumerate(years):
                value = row[4 + i].value if (4 + i) < len(row) else None
                if value is not None:
                    Measurement.objects.create(
                        country=country,
                        indicator=indicator,
                        year=year,
                        value=value
                    )
